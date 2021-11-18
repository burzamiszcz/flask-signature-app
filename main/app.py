from flask import (Flask, 
                    render_template, 
                    request, 
                    send_file, 
                    after_this_request, 
                    redirect, 
                    g, 
                    session, 
                    make_response, 
                    Response, json, flash)
import requests
from flask.helpers import url_for
# from openpyxl.workbook import workbook
from werkzeug.utils import secure_filename
import os
from os.path import isfile, join
import pandas as pd
from openpyxl.drawing.image import Image
from datetime import datetime
from operator import itemgetter
import shutil
from openpyxl import load_workbook
import PIL
import jpype
import asposecells
jpype.startJVM()
from asposecells.api import Workbook, FileFormatType, PdfSaveOptions
# from win32com import client
# import pythoncom

def make_trans_bg(path):
    img = PIL.Image.open(path)
    img = img.convert("RGBA")
    datas = img.getdata()
    newData = []
    for item in datas:
        if item[0] == 255 and item[1] == 255 and item[2] == 255:
            newData.append((255, 255, 255, 0))
        else:
            newData.append(item)
    img.putdata(newData)
    img.save(path, "PNG")

def itm_training(name_surname, path):
    make_trans_bg(f'./static/training_lists_signed/signatures/{name_surname}.png')
    signature = Image(f'./static/training_lists_signed/signatures/{name_surname}.png')
    signature.width = 95
    signature.height = 32
    name_surname = name_surname.split()
    workbook = load_workbook(path)
    sheet = workbook.active
    i = 0
    while sheet[f'B{9 + i}'].value:
        i += 1
    for j in range(len(full_data_list)):
        if name_surname[0] == full_data_list[j][1] and name_surname[1] == full_data_list[j][2]:
            sheet[f'B{9 + i}'] = str(name_surname[1] + ' ' + name_surname[0])
            sheet[f'C{9 + i}'] = full_data_list[j][4]
            sheet[f'D{9 + i}'] = full_data_list[j][5]
            sheet[f'E{9 + i}'] = today_date()
            sheet.add_image(signature, f'F{9 + i}')
            break
    for person in people_list:
        if name_surname[0] == person[0] and name_surname[1] == person[1]:
            people_list.remove(person)
            break
    workbook.save(path)

def make_training(title, teacher, time, date, agenda):
    workbook = load_workbook("./static/training_lists_empty/training_generator_template.xlsx")
    sheet = workbook.active
    sheet[f'A2'] = str(title)
    sheet[f'A4'] = str(agenda)
    sheet[f'A6'] = str(date)
    sheet[f'D6'] = str(time)
    sheet[f'E6'] = str(teacher)
    workbook.save(f"./static/training_lists_empty/{title}.xlsx")
    xlsx2pdf(f"./static/training_lists_empty/{title}.xlsx", 
            f"./static/pdf_trainings/{title}.pdf")

def group_read(path):
    sheet = pd.read_excel(path)
    list = sheet.values.tolist()
    list_only_namesurname = list[:]
    for i in range(len(list)):
        list_only_namesurname[i] = [list_only_namesurname[i][1], list_only_namesurname[i][2]]
    list_only_namesurname = sorted(list_only_namesurname, key=itemgetter(0))

    return list, list_only_namesurname

def files_in_the_path(path):
    files = [f for f in os.listdir(path) if isfile(join(path, f))]
    for i in range(len(files)):
        files[i] = files[i].split('.')[0]
    return files

def trainig_copy(src, dst):
    shutil.copy2(src, dst)

def today_date():
    now = datetime.now()
    dt_string = now.strftime("%d-%m-%Y")
    return dt_string

def xlsx2pdf(path_in, path_out):
    path_in = os.path.join(fileDir, path_in)
    path_out = os.path.join(fileDir, path_out)
    workbook = Workbook(path_in)
    saveOptions = PdfSaveOptions()
    saveOptions.setOnePagePerSheet(True)
    workbook.save(path_out, saveOptions)

    # pythoncom.CoInitialize()
    # xlApp = client.Dispatch("Excel.Application")
    # books = xlApp.Workbooks.Open(path_in)
    # ws = books.Worksheets[0]
    # ws.ExportAsFixedFormat(0, path_out)
    # books.close

def remove_files_from_folder(path_to_folder):
    for f in os.listdir(path_to_folder):
        os.remove(os.path.join(path_to_folder, f))
    

app = Flask(__name__)
app.secret_key = 'somesecretkeythatonlyishouldknow'

people_list, full_data_list = [], []
fileDir = os.path.dirname(os.path.realpath('__file__'))
UPLOAD_FOLDER = './static/groups'
ALLOWED_EXTENSIONS = {'pdf', 'xlsx'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.before_request
def before_request():
    if 'training_select' in session:
        g.training_select = session['training_select']
    else:
        g.training_select = None

@app.route('/', methods=['GET', 'POST'])
def login():
    session.clear()
    if request.method == 'POST':
        session.pop('credential', None)
        username = request.form['username'].lower()
        password = request.form['password']
        respose = requests.post("http://login:11000/", json = {'username': username, "password": password, 'email': None})
        print(respose, flush=True)

        if respose.json()['credential'] != None:
            session['credential'] = respose.json()['credential']
            session['id'] = respose.json()['id']
            session['username'] = respose.json()['username']
            print('lol', flush=True)
            return redirect(url_for('main'))
        flash("Błedny login lub hasło")
    return render_template('login.html')

@app.route("/main", methods=['POST', 'GET'])
def main():
    trainings = files_in_the_path("./static/pdf_trainings")
    groups = files_in_the_path("./static/groups")

    if request.method == 'POST':
        training_select = request.form['training_select']
        group_select = request.form['group_select']
        session['training_select'] = str(training_select)
        session['group_select'] = str(group_select)
        
        trainig_copy(f"./static/training_lists_empty/{session['training_select']}.xlsx", f"./static/training_lists_signed/{session['training_select']}-{today_date()}.xlsx")


        global full_data_list, people_list
        full_data_list, people_list = group_read("./static/groups/%s.xlsx"%session['group_select'])
        
        return redirect(url_for('training'))
        
    return render_template('main.html', trainings = trainings, groups = groups)

@app.route('/training', methods=['POST', 'GET'])
def training():
    is_list_empty = False
    training_select = f"static/pdf_trainings/{session['training_select']}.pdf"
    if request.method == 'POST':
        if request.form.get('person_select'):
            name_surname = request.form['person_select']
            session['name_surname'] = name_surname
            itm_training(name_surname, f"./static/training_lists_signed/{session['training_select']}-{today_date()}.xlsx")

        elif request.form.get('preview'):
            xlsx2pdf(f"./static/training_lists_signed/{session['training_select']}-{today_date()}.xlsx", f"./static/training_lists_signed/done/{session['training_select']}-{today_date()}.pdf")
            training_select = f"static/training_lists_signed/done/{session['training_select']}-{today_date()}.pdf"

        else:
            xlsx2pdf(f"./static/training_lists_signed/{session['training_select']}-{today_date()}.xlsx", f"./static/training_lists_signed/done/{session['training_select']}-{today_date()}.pdf")
            remove_files_from_folder("./static/training_lists_signed/signatures")
            shutil.move(f"./static/training_lists_signed/{session['training_select']}-{today_date()}.xlsx", f"./static/training_lists_signed/done/{session['training_select']}-{today_date()}.xlsx")
            return redirect(url_for('main'))

        if people_list == []:
            is_list_empty = True            
        else:
            pass
        
    return render_template('training.html', people_list = people_list, training_select = training_select, is_list_empty = is_list_empty)

@app.route("/groups", methods=['POST', 'GET'])
def groups():
    edit=False
    groups = files_in_the_path("./static/groups")
    if request.method == 'POST':
        if request.form.get('delete_group'):
            edit = True
            delete_group = request.form['delete_group']
            os.remove(f"./static/groups/{delete_group}.xlsx")
            groups = files_in_the_path("./static/groups")
        else:
            group = request.files['group']
            group.save(os.path.join(app.config['UPLOAD_FOLDER'], secure_filename('grupa-%s.xlsx'%str(today_date()))))
            groups = files_in_the_path("./static/groups")
    return render_template('groups.html', groups = groups, edit=edit)

@app.route("/add_training", methods=['POST', 'GET'])
def add_training():
    trainings = files_in_the_path("./static/pdf_trainings")
    if request.method == 'POST':
        print(request.form)
        title = request.form['title']
        teacher = request.form['teacher']
        time = request.form['time']
        date = request.form['date']
        agenda = request.form['agenda']
        make_training(title, teacher, time, date, agenda)
        
    return render_template('add_training.html', trainings = trainings)

@app.route("/signature", methods=['POST'])
def signature():
    signature_string = request.form['data']
    signature_name = request.form['name']
    signature_string = signature_string.split(',')[1]
    import base64
    imgdata = base64.b64decode(signature_string)
    filename = f'static/training_lists_signed/signatures/{signature_name}.png' 
    with open(filename, 'wb') as f:
        f.write(imgdata)
    return "Hello" #json.dumps({'status':'OK','data': signature_string})

@app.route('/add_user', methods=['GET', 'POST'])
def menu_add_user():

    if request.method == 'POST':
        username = request.form['username_add'].lower()
        password = request.form['password_add']
        password2 = request.form['password_add2']
        credential = request.form['credentials_select']
        if credential == 'Użytkownik':
            credential = 'user'
        elif credential == 'Administrator':
            credential = 'admin'
        elif credential == 'Operator':
            credential = 'operator'

        if password != password2:
            flash("Hasła nie są jednakowe")
        else:
            response = requests.post("http://login:11000/new_user", json = {"username": username, "password": password, "credential": credential})
            if response.json()['status'] == "exist":
                flash("Taki użytkownik już istnieje")
            if response.json()['status'] == "created":
                flash("Pomyślnie utworzono użytkownika")
    return render_template('add_user.html')


if __name__ == '__main__':
    app.run()
