from flask import Flask, jsonify, request
import requests
import sqlite3
import os
from openpyxl import load_workbook


from flask import (Flask, 
                    render_template, 
                    request, 
                    send_file, 
                    after_this_request, 
                    redirect, 
                    g, 
                    session, 
                    make_response, 
                    Response, json, flash)
import requests
from flask.helpers import url_for
# from openpyxl.workbook import workbook
from werkzeug.utils import secure_filename
import os
from os.path import isfile, join
import pandas as pd
from openpyxl.drawing.image import Image
from datetime import datetime
from operator import itemgetter
import shutil
from openpyxl import load_workbook
import PIL
import jpype
import asposecells
jpype.startJVM()
from asposecells.api import Workbook, FileFormatType, PdfSaveOptions

def files_in_the_path(path):
    files = [f for f in os.listdir(path) if isfile(join(path, f))]
    for i in range(len(files)):
        files[i] = files[i].split('.')[0]
    return files


fileDir = os.path.dirname(os.path.realpath('__file__'))

app = Flask(__name__)

def dict_factory(cursor, row):
    d = {}
    for idx, col in enumerate(cursor.description):
        d[col[0]] = row[idx]
    return d

def xlsx2pdf(path_in, path_out):
    path_in = os.path.join(fileDir, path_in)
    path_out = os.path.join(fileDir, path_out)
    workbook = Workbook(path_in)
    saveOptions = PdfSaveOptions()
    saveOptions.setOnePagePerSheet(True)
    workbook.save(path_out, saveOptions)


@app.route('/new_group', methods=['POST', 'GET'])
def new_group():
    group = request.files['group']
    group.save(os.path.join(app.config['UPLOAD_FOLDER'], secure_filename('grupa-%s.xlsx'%str(today_date()))))
    groups = files_in_the_path("./static/groups")


@app.route('/new_doc', methods=['POST', 'GET'])
def new_doc():

    title = request.json['title']
    teacher = request.json['teacher']
    time = request.json['time']
    date = request.json['date']
    agenda = request.json['agenda']

    workbook = load_workbook("./static/training_lists_empty/training_generator_template.xlsx")
    sheet = workbook.active
    sheet[f'A2'] = str(title)
    sheet[f'A4'] = str(agenda)
    sheet[f'A6'] = str(date)
    sheet[f'C6'] = str(time)
    sheet[f'D6'] = str(teacher)
    workbook.save(f"./static/training_lists_empty/{title}.xlsx")
    xlsx2pdf(f"./static/training_lists_empty/{title}.xlsx", 
            f"./static/pdf_trainings/{title}.pdf")

@app.route("/signature", methods=['POST'])
def signature():
    signature_string = request.form['data']
    signature_name = request.form['name']
    signature_string = signature_string.split(',')[1]
    import base64
    imgdata = base64.b64decode(signature_string)
    filename = f'static/training_lists_signed/signatures/{signature_name}.png' 
    with open(filename, 'wb') as f:
        f.write(imgdata)
    return "Hello" #json.dumps({'status':'OK','data': signature_string})

if __name__ == '__main__':
    app.run()
